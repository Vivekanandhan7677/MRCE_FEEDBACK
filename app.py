import os
from flask import Flask, render_template, request, send_file, jsonify, redirect, url_for, session
import os
import psycopg2
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4


app = Flask(__name__)
app.secret_key = "feedback_secret_key"


# ---------- DB CONNECTION ----------
#---------- FIXED: Use DATABASE_URL env variable for PostgreSQL connection ----------

DATABASE_URL = os.environ.get("DATABASE_URL")

def db():
    if not DATABASE_URL:
        raise Exception("DATABASE_URL not set")
    return psycopg2.connect(DATABASE_URL, sslmode="require")

@app.route('/favicon.ico')
def favicon():
    return send_file(os.path.join(app.root_path, 'static', 'logo.png'))
# ---------- LOAD SUBJECTS ----------
@app.route("/get_subjects")
def get_subjects():
    branch = request.args.get("branch")
    year = request.args.get("year")
    semester = request.args.get("semester")
    section = request.args.get("section")

    con = db()
    cur = con.cursor()
    cur.execute("""
        SELECT subject_name, faculty_name
        FROM subjects
        WHERE branch=%s AND year=%s AND semester=%s AND section=%s
    """, (branch, year, semester, section))
    rows = cur.fetchall()
    con.close()

    return jsonify({
        "subjects": [{"subject": r[0], "faculty": r[1]} for r in rows]
    })


# ---------- STUDENT LOGIN ----------
@app.route("/", methods=["GET", "POST"])
def student_login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        con = db()
        cur = con.cursor()

        # ✅ AUTO CREATE STUDENT
        cur.execute("SELECT role FROM users WHERE username=%s", (username,))
        user = cur.fetchone()

        if not user:
            cur.execute(
                "INSERT INTO users (username, password, role) VALUES (%s,%s,'student')",
                (username, username)
            )
            con.commit()

        # ✅ VALIDATE LOGIN
        cur.execute(
            "SELECT role FROM users WHERE username=%s AND password=%s",
            (username, password)
        )
        user = cur.fetchone()

        # ❌ INVALID LOGIN
        if not user or user[0] != "student":
            con.close()
            return "Invalid Student Login"

        # ✅ CHECK IF FEEDBACK ALREADY SUBMITTED
        cur.execute(
            "SELECT 1 FROM students_feedback WHERE roll=%s LIMIT 1",
            (username,)
        )
        feedback_exists = cur.fetchone()

        if feedback_exists:
            con.close()
            return "❌ You have already submitted feedback!"

        # ✅ ALLOW LOGIN
        session["user"] = username
        con.close()
        return redirect(url_for("student_page"))

    return render_template("login_student.html")


# ---------- STUDENT PAGE ----------
@app.route("/student", methods=["GET", "POST"])
def student_page():

    if "user" not in session:
        return redirect(url_for("student_login"))

    if request.method == "POST":

        import json

        con = db()
        cur = con.cursor()

        name = request.form.get("name")
        roll = session["user"]
        year = request.form.get("year")
        semester = request.form.get("semester")
        branch = request.form.get("branch")
        section = request.form.get("section")

        feedback_json = request.form.get("all_feedback")

        if not feedback_json:
            return "<h3>No feedback received</h3>"

        feedback_list = json.loads(feedback_json)

        for fb in feedback_list:

            subject = fb["subject"]
            suggestion = fb.get("suggestion")

            # Prevent duplicate feedback
            cur.execute("""
                SELECT id FROM students_feedback
                WHERE login_id=%s AND subject=%s
            """, (roll, subject))

            if cur.fetchone():
                continue

            # Insert feedback
            cur.execute("""
INSERT INTO students_feedback
(name, roll, year, semester, branch, section, subject, suggestion, login_id)
VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
""", (
name,
roll,
year,
semester,
branch,
section,
subject,
suggestion,
roll
))
            # Get last inserted id
            cur.execute("SELECT LASTVAL()")
            fid = cur.fetchone()[0]

            # Insert answers
            cur.execute("""
                INSERT INTO answers
                (feedback_id,q1,q2,q3,q4,q5,q6,q7,q8,q9,q10)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                fid,
                int(fb["q1"]),
                int(fb["q2"]),
                int(fb["q3"]),
                int(fb["q4"]),
                int(fb["q5"]),
                int(fb["q6"]),
                int(fb["q7"]),
                int(fb["q8"]),
                int(fb["q9"]),
                int(fb["q10"])
            ))

        con.commit()
        cur.close()
        con.close()

        return "<h2>Feedback Submitted Successfully</h2>"

    return render_template("student.html", roll=session["user"])

# ---------- LOGOUT ----------
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("student_login"))


# ---------- ADMIN LOGIN ----------
@app.route("/admin_login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        con = db()
        cur = con.cursor()
        cur.execute(
            "SELECT role FROM users WHERE username=%s AND password=%s",
            (username, password)
        )
        user = cur.fetchone()
        con.close()

        if user and user[0] == "admin":
            session["admin"] = username
            return redirect(url_for("admin"))

        return "Invalid Admin Login"

    return render_template("login_admin.html")


@app.route("/admin_logout")
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))


# ---------- ADMIN DASHBOARD ----------
@app.route("/admin", methods=["GET", "POST"])
def admin():
    if "admin" not in session:
        return redirect(url_for("admin_login"))

    percentage = 0
    responses = 0
    students = []

    if request.method == "POST":
        year = request.form["year"]
        semester = request.form["semester"]
        branch = request.form["branch"]
        section = request.form["section"]
        subject = request.form["subject"]

        con = db()
        cur = con.cursor()

        cur.execute("""
            SELECT a.q1,a.q2,a.q3,a.q4,a.q5,
                   a.q6,a.q7,a.q8,a.q9,a.q10,
                   s.suggestion,
                   s.login_id
            FROM answers a
            JOIN students_feedback s ON a.feedback_id = s.id
            WHERE s.year=%s AND s.semester=%s
            AND s.branch=%s AND s.section=%s
            AND s.subject=%s
        """, (year, semester, branch, section, subject))

        rows = cur.fetchall()
        responses = len(rows)
        suggestions = []

        if responses > 0:
            total = sum(sum(r[:10]) for r in rows)
            percentage = (total / (50 * responses)) * 100
            suggestions = [r[10] for r in rows if r[10]]
            students = [r[11] for r in rows]

        cur.execute("""
            SELECT faculty_name FROM subjects
            WHERE branch=%s AND year=%s AND semester=%s
            AND section=%s AND subject_name=%s
        """, (branch, year, semester, section, subject))

        row = cur.fetchone()
        faculty = row[0] if row else ""
        con.close()

        # Save for PDF
        session["report_data"] = {
            "rows": [r[:10] for r in rows],
            "info": (year, semester, branch, section, subject, faculty),
            "responses": responses,
            "percentage": percentage,
            "suggestions": suggestions
        }

    return render_template("admin.html",
                           percentage=percentage,
                           responses=responses,
                           students=students)

@app.route("/report")
def report():
    data = session.get("report_data")
    if not data:
        return "Generate report first"

    rows = data["rows"]
    year, semester, branch, section, subject, faculty = data["info"]
    responses = data["responses"]
    suggestions = data.get("suggestions", [])
    percentage = data["percentage"]

    # -------- Count ratings --------
    counts = [[0]*5 for _ in range(10)]
    for row in rows:
        for i, val in enumerate(row):
            counts[i][val-1] += 1

    file = "feedback_report.pdf"
    doc = SimpleDocTemplate(
        file,
        pagesize=A4,
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    styles = getSampleStyleSheet()
    elements = []

    # ---------- HEADER ----------
    logo_path = os.path.join(app.root_path, 'static', 'logo.png')
    img = Image(logo_path, width=55, height=55)
    img.hAlign = 'CENTER'
    elements.append(img)
    elements.append(Spacer(1, 6))

    elements.append(Paragraph(
        "<b><font size=18 color='blue'>MALLA REDDY COLLEGE OF ENGINEERING</font></b>",
        styles['Title']
    ))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph("<b>Feedback Form Report</b>", styles['Heading2']))
    elements.append(Spacer(1, 12))

    # ---------- INFO TABLE ----------
    info = [
        ["Branch", branch, "Year", year, "Semester", semester],
        ["Section", section, "Subject", subject, "Faculty", faculty],
        ["Responses", str(responses), "", "", ""]
    ]

    t = Table(info, colWidths=[55,50,55,180,100])
    t.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]))

    elements.append(t)
    elements.append(Spacer(1, 12))

    # ---------- PERCENTAGE ----------
    elements.append(Paragraph(
        f"<b>Overall Feedback Percentage: {percentage:.2f}%</b>",
        styles['Normal']
    ))
    elements.append(Spacer(1, 12))

    # ---------- QUESTIONS ----------
    questions = [
        "Teacher comes to the class in time",
        "Teacher comes well planned and prepared",
        "Teacher speaks clearly and audibly",
        "Teacher provides examples effectively",
        "Use of ICT tools while teaching",
        "Teacher encourages doubts and answers well",
        "Teacher is courteous and impartial",
        "Teacher maintains discipline",
        "Teacher completes syllabus at proper pace",
        "Teacher gives feedback on answer scripts"
    ]

    # Header row
    table_data = [[
        "Sl.No",
        Paragraph("<b>Criteria</b>", styles['Normal']),
        "VeryGood(5)",
        "Good(4)",
        "Average(3)",
        "BelowAvg(2)",
        "Poor(1)",
        "Total"
    ]]

    grand_total = [0]*5
    overall_total = 0

    # Question rows
    for i, q in enumerate(questions):
        total = sum(counts[i])
        overall_total += total

        for j in range(5):
            grand_total[j] += counts[i][j]

        table_data.append([
            i+1,
            Paragraph(q, styles['Normal']),
            counts[i][4],
            counts[i][3],
            counts[i][2],
            counts[i][1],
            counts[i][0],
            total
        ])

    # Grand Total Row
    table_data.append([
        "",
        Paragraph("<b>Grand Total</b>", styles['Normal']),
        grand_total[4],
        grand_total[3],
        grand_total[2],
        grand_total[1],
        grand_total[0],
        overall_total
    ])

    table = Table(table_data, colWidths=[35,165,55,55,55,65,55,50])

    table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
        ('ALIGN', (2,1), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 15))

    # ---------- SUGGESTIONS ----------
    if suggestions:
        elements.append(Paragraph("<b>Student Suggestions</b>", styles['Heading3']))
        elements.append(Spacer(1, 6))
        for i, s in enumerate(suggestions, 1):
            elements.append(Paragraph(f"{i}. {s}", styles['Normal']))
            elements.append(Spacer(1, 4))

    elements.append(Spacer(1, 70))

    # ---------- SIGNATURE ----------
    sign_table = Table(
        [["HOD Signature", "", "Principal Signature"],
         ["", "", ""]],
        colWidths=[220, 80, 220]
    )

    sign_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('LINEABOVE', (0,1), (0,1), 1, colors.black),
        ('LINEABOVE', (2,1), (2,1), 1, colors.black),
    ]))

    elements.append(sign_table)

    doc.build(elements)
    return send_file(file, as_attachment=True)






# ---------- RESET ----------
@app.route("/reset_feedback")
def reset_feedback():
    con = db()
    cur = con.cursor()

    cur.execute("TRUNCATE TABLE answers RESTART IDENTITY CASCADE")
    cur.execute("TRUNCATE TABLE students_feedback RESTART IDENTITY CASCADE")

    con.commit()
    con.close()

    return "All Feedback Reset Successfully!"

from flask import request, render_template
import psycopg2

@app.route('/admin_report', methods=['GET', 'POST'])
def admin_report():
    data = []
    summary = []

    # 👉 store selected values
    selected = {
        "branch": "",
        "year": "",
        "semester": "",
        "section": ""
    }

    if request.method == 'POST':
        branch = request.form['branch']
        year = request.form['year']
        semester = request.form['semester']
        section = request.form['section']

        # ✅ save selected values
        selected = {
            "branch": branch,
            "year": year,
            "semester": semester,
            "section": section
        }

        conn = db()
        cur = conn.cursor()

        # 1. Student + Answers
        cur.execute("""
        SELECT sf.name, sf.roll, sf.subject,
               a.q1, a.q2, a.q3, a.q4, a.q5,
               a.q6, a.q7, a.q8, a.q9, a.q10
        FROM students_feedback sf
        JOIN answers a ON sf.id = a.feedback_id
        WHERE sf.branch=%s AND sf.year=%s 
        AND sf.semester=%s AND sf.section=%s
        """, (branch, year, semester, section))

        data = cur.fetchall()

        # 2. Subject Summary
        cur.execute("""
        SELECT subject, COUNT(*) 
        FROM students_feedback
        WHERE branch=%s AND year=%s 
        AND semester=%s AND section=%s
        GROUP BY subject
        """, (branch, year, semester, section))

        summary = cur.fetchall()

        cur.close()
        conn.close()

    return render_template(
        "admin_report.html",
        data=data,
        summary=summary,
        selected=selected   # ✅ pass to HTML
    )
from flask import send_file, request
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
import io, os

@app.route('/download_excel', methods=['POST'])
def download_excel():
    branch = request.form['branch']
    year = request.form['year']
    semester = request.form['semester']
    section = request.form['section']

    conn = db()
    cur = conn.cursor()

    cur.execute("""
    SELECT sf.name, sf.roll, sf.subject,
           a.q1, a.q2, a.q3, a.q4, a.q5,
           a.q6, a.q7, a.q8, a.q9, a.q10
    FROM students_feedback sf
    JOIN answers a ON sf.id = a.feedback_id
    WHERE sf.branch=%s AND sf.year=%s 
    AND sf.semester=%s AND sf.section=%s
    """, (branch, year, semester, section))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    # ✅ Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback Report"

    # ---------- LOGO ----------
    logo_path = os.path.join(app.root_path, 'static', 'logo.png')
    if os.path.exists(logo_path):
        logo = XLImage(logo_path)
        logo.width = 80
        logo.height = 80
        ws.add_image(logo, "A1")

    # ---------- COLLEGE NAME ----------
    ws.merge_cells("A5:M5")
    cell = ws["A5"]
    cell.value = "MALLA REDDY COLLEGE OF ENGINEERING"
    cell.font = Font(size=16, bold=True)
    cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A6:M6")
    ws["A6"] = "Feedback Report"
    ws["A6"].alignment = Alignment(horizontal="center")

    # ---------- TABLE HEADER ----------
    headers = ["Name", "Roll", "Subject",
               "Q1","Q2","Q3","Q4","Q5",
               "Q6","Q7","Q8","Q9","Q10"]

    ws.append([])
    ws.append(headers)

    for col in ws[8]:
        col.font = Font(bold=True)
        col.alignment = Alignment(horizontal="center")

    # ---------- DATA ----------
    for row in rows:
        ws.append(row)

    # ---------- ⭐ RATING SCALE TABLE ----------
    start_row = ws.max_row + 3

    ws[f"A{start_row}"] = "Rating Scale"
    ws[f"A{start_row}"].font = Font(bold=True)

    rating_data = [
        ("Rating", "Meaning"),
        (5, "Very Good"),
        (4, "Good"),
        (3, "Average"),
        (2, "Below Average"),
        (1, "Poor")
    ]

    for i, row in enumerate(rating_data, start=start_row + 1):
        ws[f"A{i}"] = row[0]
        ws[f"B{i}"] = row[1]

        ws[f"A{i}"].alignment = Alignment(horizontal="center")
        ws[f"B{i}"].alignment = Alignment(horizontal="center")

        if i == start_row + 1:  # header row
            ws[f"A{i}"].font = Font(bold=True)
            ws[f"B{i}"].font = Font(bold=True)

    # ---------- AUTO WIDTH ----------
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # ---------- SIGNATURES ----------
    last_row = ws.max_row + 4

    ws.merge_cells(f"A{last_row}:C{last_row}")
    ws.merge_cells(f"K{last_row}:M{last_row}")

    ws[f"A{last_row}"] = "HOD Signature"
    ws[f"K{last_row}"] = "Principal Signature"

    ws[f"A{last_row}"].alignment = Alignment(horizontal="center")
    ws[f"K{last_row}"].alignment = Alignment(horizontal="center")

    # ---------- SIGNATURE IMAGES ----------
    hod_sign = os.path.join(app.root_path, 'static', 'hod_sign.png')
    principal_sign = os.path.join(app.root_path, 'static', 'principal_sign.png')

    if os.path.exists(hod_sign):
        img1 = XLImage(hod_sign)
        img1.width = 120
        img1.height = 50
        ws.add_image(img1, f"A{last_row+1}")

    if os.path.exists(principal_sign):
        img2 = XLImage(principal_sign)
        img2.width = 120
        img2.height = 50
        ws.add_image(img2, f"K{last_row+1}")

    # ---------- SAVE ----------
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                     download_name="feedback_report.xlsx",
                     as_attachment=True)
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)